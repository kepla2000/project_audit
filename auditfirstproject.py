from tkinter import *
from tkinter import filedialog
from tkinter import messagebox
import os
import sqlite3
import pandas as pd
from sqlalchemy import create_engine
import tabula
from PyPDF2 import PdfFileReader 
import shutil
from tkinter import ttk
from docx2pdf import convert
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import numpy as np
import re 
import aspose.words as aw





#Main Window Creation

root = Tk()
root.title("Audit X")
root.geometry("800x300")




pd.set_option("display.max_rows",None)
pd.set_option("display.max_columns",None)






#function for analysis and report of an xlsx file

def report_xlsx():
	#Everything about the current file

	global ws
	global ws1
	#finding the number and the name of the courses of that particular year(current file)
	wb = load_workbook(root_filename)
	ws = wb["Sheet2"]
	
	number_of_courses = 0
	none_counter = 0
	for i in range(20,100):
		if ws.cell(row = 14 , column = i).value != "None":
			number_of_courses = number_of_courses + 1
		else:
			break
			

	#range of cells for the courses
	courses = []
	for i in range(20, 20 + number_of_courses):
		courses.append(ws.cell(row = 14, column = i).value)

	courses_clean = filter(None, courses)
	courses_clean1 = list(courses_clean)


	#finding the the number of courses for (current) semester
	noc1 = len(courses_clean1)

	#finding the number of students that passed the semester current(numberofrows_current)
	numberofrows = 0
	for i in range(15,3000):
		if ws.cell(row=i , column=3).value != None:
			numberofrows = numberofrows + 1
		else:
			break

	#finding the cell that contains the value of the number of the number of students trailing(current)
	N = (15 + numberofrows) + 2

	#finding the rows number of students that are trailing up to four courses(current)
	numberofrows_trailing4 = 0
	for i in range(N + 5,3000):
		if ws.cell(row=i , column=4).value != None:
			numberofrows_trailing4 = numberofrows_trailing4 + 1
		else:
			break



	#Everything about the previous file


	#finding the number of courses for the particular year(previous file)
	wb1 = load_workbook(root_filename1)
	ws1 = wb1["Sheet2"]

	number_of_courses1 = 0
	none_counter1 = 0
	for i in range(20,100):
		if ws1.cell(row = 14 , column = i).value != "None":
			number_of_courses1 = number_of_courses1 + 1
		else:
			break


	courses1 = []
	for i in range(20, 20 + number_of_courses1):
		courses1.append(ws1.cell(row = 14, column = i).value)

	courses_clean11 = filter(None, courses1)
	courses_clean2 = list(courses_clean11)

	#list of courses (previous)

	#finding the number of courses for previous semester
	noc2 = len(courses_clean2)

	numberofrows_previous = 0
	for i in range(15,3000):
		if ws1.cell(row=i , column=3).value != None:
			numberofrows_previous = numberofrows_previous + 1
		else:
			break
	#finding the row that contains the number of students trailing up to four courses(previous)

	#finding the number of students that are trailing up to four courses(current)
	N_previous= (15 + numberofrows_previous) + 2
	
	#finding the rows number of students that are trailing up to four courses(previous)
	numberofrows_trailing4_previous = 0
	for i in range(N_previous + 5,3000):
		if ws.cell(row=i , column=4).value != None:
			numberofrows_trailing4_previous = numberofrows_trailing4_previous + 1
		else:
			break






	#system file related code

	#deleting existing files
	if os.path.exists("original_database.db"):
		os.remove("original_database.db")







	#making a connection to a database(original_database.db)
	connection = sqlite3.connect("original_database.db")
	c = connection.cursor()


	#finding the number of students that passe in the current semester and placing in a database
	global engine
	engine = create_engine('sqlite:///original_database.db')
	df = pd.read_excel(root_filename,sheet_name = 1 ,header=12, usecols='C:AH', skiprows = 0,nrows = numberofrows + 1)
	del df["Unnamed: 4"]
	del df["Unnamed: 18"]
	del df["Unnamed: 12"]

	df.to_sql("""NUMBER_OF_STUDENTS_PASSED_CURRENT""", con = engine, if_exists = "replace",  index = False )
	
	
	#finding the number of students that passed previous and placing in a database(previous)

	df1 = pd.read_excel(root_filename1, sheet_name = 1, header = 12, usecols="C:AH", skiprows = 0,nrows = numberofrows_previous + 1)
	del df1["Unnamed: 4"]
	del df1["Unnamed: 18"]
	del df1["Unnamed: 12"]

	df1.to_sql("NUMBER_OF_STUDENTS_PASSED_PREVIOUS", con = engine, if_exists = "replace", index = False )
	global df2
	global df3

	#finding the number of students that are trailing and more(current)(if they exists)
	df2 = pd.read_excel(root_filename, sheet_name = 1 , header = 12 , usecols = "C:AH",skiprows = 0)
	df3 = pd.read_excel(root_filename1, sheet_name = 1 , header = 12, usecols = "C:AH", skiprows = 0)
	global header1_1 
	global header2_2 
	global header3_3 
	global header4_4
	global header1_11
	global header2_22 
	global header3_33 
	global header4_44
	

	header1_1 = ""
	header2_2 = ""
	header3_3 = ""
	header4_4 = ""
	header1_11 = ""
	header2_22 = ""
	header3_33 = ""
	header4_44 = ""

	for i in range(0, len(df2)):
		if df2.iat[i , 0] == "The following students are trailing up to four courses":
			header1_1 = i
			students_trailing_up_to_four_courses()
		elif df2.iat[i , 0] == "The following students are trailing more than four courses":
			header2_2 = i
			students_trailing_more_than_four_courses()
		elif df2.iat[i, 0] == "The following students have CWA less than 45:":
			header3_3 = i
			students_with_CWA_less_than_45()
		elif df2.iat[i, 0] == "The following students have been WITHDRAWN for the reason(s) stated below:":
			header4_4 = i 
			Abandoned()

	for i in range(0, len(df3)):
		if df3.iat[i, 0] == "The following students are trailing up to four courses":
			header1_11 = i
			students_trailing_up_to_four_coursesp()
		elif df3.iat[i , 0] == "The following students are trailing more than four courses":
			header2_22 = i
			students_trailing_more_than_four_coursesp()
		elif df3.iat[i, 0] == "The following students have CWA less than 45:":
			header3_33 = i
			students_with_CWA_less_than_45p()
		elif df3.iat[i, 0] == "The following students have been WITHDRAWN for the reason(s) stated below:":
			header4_44 = i 
			Abandonedp()

	xlsx_audit_function_2FS()
	auditing_function2_checking_credit_registered2FS()







def report_xlsx1():
	wb = load_workbook(root_filename)
	ws = wb["Sheet2"]

	wb1 = load_workbook(root_filename1)
	ws1 = wb1["Sheet2"]

	wb2 = load_workbook(root_filename2)
	ws2 = wb2["Sheet2"]


	numberofrowsTFS = 0
	for i in range(15,3000):
		if ws.cell(row = i , column=3).value != None:
			numberofrowsTFS = numberofrowsTFS + 1
		else:
			break



	numberofrows_previousTFS = 0
	for i in range(15,3000):
		if ws1.cell(row=i , column=3).value != None:
			numberofrows_previousTFS = numberofrows_previousTFS + 1
		else:
			break

	numberofrows_supplementaryTFS = 0
	for i in range(15, 3000):
		if ws2.cell(row = i, column = 3).value != None:
			numberofrows_supplementaryTFS = numberofrows_supplementaryTFS + 1
		else:
			break

		

	global header111,header222,header333,header444,header111_1,header222_2,header333_3,header444_4,header111_11,header222_22,header333_33,header444_44
	global engine1

	connection11 = sqlite3.connect("original_database2.db")
	c1 = connection11.cursor()

	engine1 = create_engine("sqlite:///original_database2.db")

	#for the current file
	df4 = pd.read_excel(root_filename, sheet_name = 1, header = 12, usecols = "C:AH", skiprows = 0, nrows = numberofrowsTFS + 1)
	del df4["Unnamed: 4"]
	del df4["Unnamed: 18"]
	del df4["Unnamed: 12"]

	df4.to_sql("""NUMBER_OF_STUDENTS_PASSED_CURRENT_TFS""", con = engine1, if_exists = "replace",  index = False )




	#for the previous file
	df5 = pd.read_excel(root_filename1, sheet_name =  1, header = 12, usecols = "C:AH",skiprows = 0, nrows = numberofrows_previousTFS + 1)
	del df5["Unnamed: 4"]
	del df5["Unnamed: 18"]
	del df5["Unnamed: 12"]

	df5.to_sql("NUMBER_OF_STUDENTS_PASSED_PREVIOUS_TFS", con = engine1, if_exists = "replace", index = False)



	df6 = pd.read_excel(root_filename2, sheet_name = 1, header = 12, usecols = "C:AH", skiprows = 0, nrows = numberofrows_supplementaryTFS + 1)
	del df6["Unnamed: 4"]
	del df6["Unnamed: 18"]
	del df6["Unnamed: 12"]


	df6.to_sql("NUMBER_OF_STUDENTS_PASSED_SUPPLEMENTARY_TFS",con = engine1, if_exists = "replace", index = False)
	global df7
	global df9
	global df8
	#finding the number of students that are trailing and more
	df7 = pd.read_excel(root_filename, sheet_name = 1, header = 12, usecols ="C:AH", skiprows = 0)
	df8 = pd.read_excel(root_filename1, sheet_name = 1, header = 12, usecols = "C:AH", skiprows = 0)
	df9 = pd.read_excel(root_filename2, sheet_name = 1, header = 12, usecols = "C:AH", skiprows = 0)

	header111 = ""
	header222 = ""
	header333 = ""
	header444 = ""
	header111_1 = ""
	header222_2 = ""
	header333_3 = ""
	header444_4 = ""
	header111_11 = ""
	header222_22 = ""
	header333_33 = ""
	header444_44 = ""
	

	for i in range(0, len(df7)):
		if df7.iat[i , 0] == "The following students are trailing up to four courses":
			header111 = i
			students_trailing_up_to_four_coursesTFS()
		elif df7.iat[i , 0] == "The following students are trailing more than four courses":
			header222 = i
			students_trailing_more_than_four_coursesTFS()
		elif df7.iat[i, 0] == "The following students have CWA less than 45:":
			header333 = i
			print("hello")
			students_with_CWA_less_than_45TFS()
		elif df7.iat[i, 0] == "The following students have been WITHDRAWN for the reason(s) stated below:":
			header444 = i 
			AbandonedTFS()


	for i in range(0, len(df8)):
		if df8.iat[i , 0] == "The following students are trailing up to four courses":
			header111_1 = i
			students_trailing_up_to_four_coursesTFSp()
		elif df8.iat[i , 0] == "The following students are trailing more than four courses":
			header222_2 = i
			students_trailing_more_than_four_coursesTFSp()
		elif df8.iat[i, 0] == "The following students have CWA less than 45:":
			header333_3 = i
			students_with_CWA_less_than_45TFSp()
		elif df8.iat[i, 0] == "The following students have been WITHDRAWN for the reason(s) stated below:":
			header444_4 = i 
			AbandonedTFSp()

	for i in range(0, len(df9)):
		if df9.iat[i , 0] == "The following students are trailing up to four courses":
			header111_11 = i
			students_trailing_up_to_four_coursesTFSs()
		elif df9.iat[i , 0] == "The following students are trailing more than four courses":
			header222_22 = i
			students_trailing_more_than_four_coursesTFSs()
		elif df9.iat[i, 0] == "The following students have CWA less than 45:":
			header333_33 = i
			print("hello")
			students_with_CWA_less_than_45TFSs()
		elif df9.iat[i, 0] == "The following students have been WITHDRAWN for the reason(s) stated below:":
			header444_44 = i 
			AbandonedTFSs()

	xlsx_audit_function_3FS()
	



#CURRNT TFS

def students_trailing_up_to_four_coursesTFS():
	rownumberTU4TFS = 0
	for i in range(header111 + 3, len(df7)):
		if pd.notnull(df7.iat[i, 1]):
			rownumberTU4TFS = rownumberTU4TFS + 1
		else:
			break
	new_df_7 = df7.iloc[range(header111 + 1, header111 + 2 + rownumberTU4TFS + 1),:]
	del new_df_7["Unnamed: 4"]
	del new_df_7["Unnamed: 18"]
	del new_df_7["Unnamed: 12"]
	new_df_7.to_sql("STUDENTS_TRAILING_UP_TO_FOUR_COURSES_CURRENT_TFS", con = engine1, if_exists = "replace", index = False)

def students_trailing_more_than_four_coursesTFS():
	rownumberTM4TFS = 0
	for i in range(header222 + 3, len(df7)):
		if pd.notnull(df7.iat[i, 1]):
			rownumberTM4TFS = rownumberTM4TFS + 1
		else:
			break
	new_df_7 = df7.iloc[range(header222 + 1, header222 + 2 + rownumberTM4TFS + 1),:]
	del new_df_7["Unnamed: 4"]
	del new_df_7["Unnamed: 18"]
	del new_df_7["Unnamed: 12"]
	new_df_7.to_sql("STUDENTS_TRAILING_MORE_THAN_FOUR_COURSES_CURRENT_TFS", con = engine1, if_exists = "replace", index = False)

def students_with_CWA_less_than_45TFS():
	rownumberCWATFS = 0
	for i in range(header333 + 3, len(df7)):
		if pd.notnull(df7.iat[i, 1]):
			rownumberCWATFS = rownumberCWATFS + 1
		else:
			break
	new_df_7 = df7.iloc[range(header333 + 1, header333 + 2 + rownumberCWATFS + 1),:]
	del new_df_7["Unnamed: 4"]
	del new_df_7["Unnamed: 18"]
	del new_df_7["Unnamed: 12"]
	new_df_7.to_sql("STUDENTS_WITH_CWA_LESS_THAN_45_CURRENT_TFS", con = engine1, if_exists = "replace", index = False)

def AbandonedTFS():
	rownumberCWATFS = 0
	for i in range(header444 + 3, len(df7)):
		if pd.notnull(df7.iat[i, 1]):
			rownumberCWATFS = rownumberCWATFS + 1
		else:
			break
	new_df_7 = df7.iloc[range(header444 + 1, header444 + 2 + rownumberCWATFS + 1),:]
	del new_df_7["Unnamed: 4"]
	del new_df_7["Unnamed: 18"]
	del new_df_7["Unnamed: 12"]
	new_df_7.to_sql("ABANDONED_COURSE_CURRENT_TFS", con = engine1, if_exists = "replace", index = False)


#PREVIOUS TFS

def students_trailing_up_to_four_coursesTFSp():
	rownumberTU4TFSp = 0
	for i in range(header111_1 + 3, len(df8)):
		if pd.notnull(df8.iat[i, 1]):
			rownumberTU4TFSp = rownumberTU4TFSp + 1
		else:
			break
	new_df_8 = df8.iloc[range(header111_1 + 1, header111_1 + 2 + rownumberTU4TFSp + 1),:]
	del new_df_8["Unnamed: 4"]
	del new_df_8["Unnamed: 18"]
	del new_df_8["Unnamed: 12"]
	new_df_8.to_sql("STUDENTS_TRAILING_UP_TO_FOUR_COURSES_PREVIOUS_TFS", con = engine1, if_exists = "replace", index = False)

def students_trailing_more_than_four_coursesTFSp():
	rownumberTM4TFSp = 0
	for i in range(header222_2 + 3, len(df8)):
		if pd.notnull(df8.iat[i, 1]):
			rownumberTM4TFSp = rownumberTM4TFSp + 1
		else:
			break
	new_df_8 = df8.iloc[range(header222_2 + 1, header222_2 + 2 + rownumberTM4TFSp + 1),:]
	del new_df_8["Unnamed: 4"]
	del new_df_8["Unnamed: 18"]
	del new_df_8["Unnamed: 12"]
	new_df_8.to_sql("STUDENTS_TRAILING_MORE_THAN_FOUR_COURSES_PREVIOUS_TFS", con = engine1, if_exists = "replace", index = False)

def students_with_CWA_less_than_45TFSp():
	rownumberCWATFSp = 0
	for i in range(header333_3 + 3, len(df8)):
		if pd.notnull(df8.iat[i, 1]):
			rownumberCWATFSp = rownumberCWATFSp + 1
		else:
			break
	new_df_8 = df8.iloc[range(header333_3 + 1, header333_3 + 2 + rownumberCWATFSp + 1),:]
	del new_df_8["Unnamed: 4"]
	del new_df_8["Unnamed: 18"]
	del new_df_8["Unnamed: 12"]
	new_df_8.to_sql("STUDENTS_WITH_CWA_LESS_THAN_45_PREVIOUS_TFS", con = engine1, if_exists = "replace", index = False)

def AbandonedTFSp():
	rownumberCWATFSp = 0
	for i in range(header444_4 + 3, len(df8)):
		if pd.notnull(df8.iat[i, 1]):
			rownumberCWATFSp = rownumberCWATFSp + 1
		else:
			break
	new_df_8 = df8.iloc[range(header444_4 + 1, header444_4 + 2 + rownumberCWATFSp + 1),:]
	del new_df_8["Unnamed: 4"]
	del new_df_8["Unnamed: 18"]
	del new_df_8["Unnamed: 12"]
	new_df_8.to_sql("ABANDONED_COURSE_PREVIOUS_TFS", con = engine1, if_exists = "replace", index = False)


#SUPPLEMENTARY TFS


def students_trailing_up_to_four_coursesTFSs():
	rownumberTU4TFSs = 0
	for i in range(header111_11 + 3, len(df9)):
		if pd.notnull(df9.iat[i, 1]):
			rownumberTU4TFSs = rownumberTU4TFSs + 1
		else:
			break
	new_df_9 = df9.iloc[range(header111_11 + 1, header111_11 + 2 + rownumberTU4TFSs + 1),:]
	del new_df_9["Unnamed: 4"]
	del new_df_9["Unnamed: 18"]
	del new_df_9["Unnamed: 12"]
	new_df_9.to_sql("STUDENTS_TRAILING_UP_TO_FOUR_COURSES_SUPPLEMENTARY_TFS", con = engine1, if_exists = "replace", index = False)

def students_trailing_more_than_four_coursesTFSs():
	rownumberTM4TFSs = 0
	for i in range(header222_22 + 3, len(df9)):
		if pd.notnull(df9.iat[i, 1]):
			rownumberTM4TFSs = rownumberTM4TFSs + 1
		else:
			break
	new_df_9 = df9.iloc[range(header222_22 + 1, header222_22 + 2 + rownumberTM4TFSs + 1),:]
	del new_df_9["Unnamed: 4"]
	del new_df_9["Unnamed: 18"]
	del new_df_9["Unnamed: 12"]
	new_df_9.to_sql("STUDENTS_TRAILING_MORE_THAN_FOUR_COURSES_SUPPLEMENTARY_TFS", con = engine1, if_exists = "replace", index = False)

def students_with_CWA_less_than_45TFSs():
	rownumberCWATFSs = 0
	for i in range(header333_33 + 3, len(df9)):
		if pd.notnull(df9.iat[i, 1]):
			rownumberCWATFSs = rownumberCWATFSs + 1
		else:
			break
	new_df_9 = df9.iloc[range(header333_33 + 1, header333_33 + 2 + rownumberCWATFSs + 1),:]
	del new_df_9["Unnamed: 4"]
	del new_df_9["Unnamed: 18"]
	del new_df_9["Unnamed: 12"]
	new_df_9.to_sql("STUDENTS_WITH_CWA_LESS_THAN_45_SUPPLEMENTARY_TFS", con = engine1, if_exists = "replace", index = False)

def AbandonedTFSs():
	rownumberCWATFSs = 0
	for i in range(header444_44 + 3, len(df9)):
		if pd.notnull(df9.iat[i, 1]):
			rownumberCWATFSs = rownumberCWATFSs + 1
		else:
			break
	new_df_9 = df9.iloc[range(header444_44 + 1, header444_44 + 2 + rownumberCWATFSs + 1),:]
	del new_df_9["Unnamed: 4"]
	del new_df_9["Unnamed: 18"]
	del new_df_9["Unnamed: 12"]
	new_df_9.to_sql("ABANDONED_COURSE_SUPPLEMENTARY_TFS", con = engine1, if_exists = "replace", index = False)











def report_pdf():
	#analysing the first semester pdf file
	#findng the number of pages of the specified pdf file
	with open(root_filename, "rb") as pdf_file:
		pdf_reader = PdfFileReader(pdf_file)
		acc_page = pdf_reader.numPages
		#creating folder for complex semester one filse (after convertion to xlsx file)
	directory = ["C:\\Users\\kepla\\Downloads\\New_folder1","C:\\Users\\kepla\\Downloads\\New_folder1\\complex_pdf_semester1(converted)","C:\\Users\\kepla\\Downloads\\New_folder1\\simple_pdf_semester1(converted)"]
	for i in directory:
			if os.path.exists(i):
				shutil.rmtree(i)
				os.mkdir(i)
				pass
			else:
				os.mkdir(i)

		
	#iterating through all the pages of the pdf file and converting it into an xlsx file and separating each page 
	#range scans values from the specified value to the end value + 1 
	for page in range(1, acc_page + 1):
		#creating a dataframe to contain information of the pdf file
		df_11 = tabula.read_pdf(root_filename, pages = page, stream = True)[0]
		df_11.to_excel(f"C:\\Users\\kepla\\Downloads\\New_folder1\\converted_semester1_{page}.xlsx")

	#moving complex files to a different folder
	for i in range(2, acc_page + 1):
		shutil.move(f"C:\\Users\\kepla\\Downloads\\New_folder1\\converted_semester1_{i}.xlsx","C:\\Users\\kepla\\Downloads\\New_folder1\\complex_pdf_semester1(converted)")

	#moving simple file to simple_pdf_semester1(converted)
	shutil.move(f"C:\\Users\\kepla\\Downloads\\New_folder1\\converted_semester1_1.xlsx","C:\\Users\\kepla\\Downloads\\New_folder1\\simple_pdf_semester1(converted)")
	#combining all the xlsx file into one master file
	current_directory = os.listdir("C:\\Users\\kepla\\Downloads\\New_folder1\\complex_pdf_semester1(converted)")
	excel_list = []

	for files in current_directory:
		excel_list.append(pd.read_excel(f"C:\\Users\\kepla\\Downloads\\New_folder1\\complex_pdf_semester1(converted)\\{files}"))

	#creating an empty dataframe
	excel_merged = pd.DataFrame()
	
	for excel_file in excel_list:
		excel_merged = excel_merged.append(excel_file, ignore_index = True)

	excel_merged.to_excel("C:\\Users\\kepla\\Downloads\\New_folder1\\complex_pdf_semester1(converted)\\final_converted_complex_result1.xlsx")

	#taking second file and analysing
	#finding the number of pages of the file
	with open(root_filename1, "rb") as pdf_file1:
		pdf_reader1 = PdfFileReader(pdf_file1)
		acc_page1 = pdf_reader1.numPages
		#creating folder for complex semester one filse (after convertion to xlsx file)
	directory1 = ["C:\\Users\\kepla\\Downloads\\New_folder2","C:\\Users\\kepla\\Downloads\\New_folder2\\complex_pdf_semester2(converted)","C:\\Users\\kepla\\Downloads\\New_folder2\\simple_pdf_semester2(converted)"]
	for i in directory1:
			if os.path.exists(i):
				shutil.rmtree(i)
				os.mkdir(i)
				pass
			else:
				os.mkdir(i)

		
	#iterating through all the pages of the pdf file and converting it into an xlsx file and separating each page 
	#range scans values from the specified value to the end value + 1 
	for page1 in range(1, acc_page1 + 1):
		#creating a dataframe to contain information of the pdf file
		df_22 = tabula.read_pdf(root_filename1, pages = page1, stream = True)[0]
		df_22.to_excel(f"C:\\Users\\kepla\\Downloads\\New_folder2\\converted_semester2_{page1}.xlsx")

	#moving complex files to a different folder
	for i in range(2, acc_page1 + 1):
		shutil.move(f"C:\\Users\\kepla\\Downloads\\New_folder2\\converted_semester2_{i}.xlsx","C:\\Users\\kepla\\Downloads\\New_folder2\\complex_pdf_semester2(converted)")

	#moving simple file to simple_pdf_semester1(converted)
	shutil.move(f"C:\\Users\\kepla\\Downloads\\New_folder2\\converted_semester2_1.xlsx","C:\\Users\\kepla\\Downloads\\New_folder2\\simple_pdf_semester2(converted)")
	#combining all the xlsx file into one master file
	current_directory1 = os.listdir("C:\\Users\\kepla\\Downloads\\New_folder2\\complex_pdf_semester2(converted)")
	excel_list1 = []

	for files1 in current_directory1:
		excel_list1.append(pd.read_excel(f"C:\\Users\\kepla\\Downloads\\New_folder2\\complex_pdf_semester2(converted)\\{files1}"))

	#creating an empty dataframe
	excel_merged1 = pd.DataFrame()
	
	for excel_file1 in excel_list1:
		excel_merged1 = excel_merged1.append(excel_file1, ignore_index = True)

	excel_merged1.to_excel("C:\\Users\\kepla\\Downloads\\New_folder2\\complex_pdf_semester2(converted)\\final_converted_complex_result2.xlsx")

	#taking supplementary file and analysing

	with open(root_filename2, "rb") as pdf_file2:
		pdf_reader2 = PdfFileReader(pdf_file2)
		acc_page2 = pdf_reader2.numPages
		#creating folder for complex semester one filse (after convertion to xlsx file)
	directory1 = ["C:\\Users\\kepla\\Downloads\\New_folder3","C:\\Users\\kepla\\Downloads\\New_folder3\\complex_pdf_semester3(converted)","C:\\Users\\kepla\\Downloads\\New_folder3\\simple_pdf_semester3(converted)"]
	for i in directory1:
			if os.path.exists(i):
				shutil.rmtree(i)
				os.mkdir(i)
				pass
			else:
				os.mkdir(i)

		
	#iterating through all the pages of the pdf file and converting it into an xlsx file and separating each page 
	#range scans values from the specified value to the end value + 1 
	for page2 in range(1, acc_page2 + 1):
		#creating a dataframe to contain information of the pdf file
		df_33 = tabula.read_pdf(root_filename1, pages = page2, stream = True)[0]
		df_33.to_excel(f"C:\\Users\\kepla\\Downloads\\New_folder3\\converted_semester3_{page2}.xlsx")

	#moving complex files to a different folder
	for i in range(2, acc_page2 + 1):
		shutil.move(f"C:\\Users\\kepla\\Downloads\\New_folder3\\converted_semester3_{i}.xlsx","C:\\Users\\kepla\\Downloads\\New_folder3\\complex_pdf_semester3(converted)")

	#moving simple file to simple_pdf_semester1(converted)
	shutil.move(f"C:\\Users\\kepla\\Downloads\\New_folder3\\converted_semester3_1.xlsx","C:\\Users\\kepla\\Downloads\\New_folder3\\simple_pdf_semester3(converted)")
	#combining all the xlsx file into one master file
	current_directory2 = os.listdir("C:\\Users\\kepla\\Downloads\\New_folder3\\complex_pdf_semester3(converted)")
	excel_list2 = []

	for files2 in current_directory2:
		excel_list1.append(pd.read_excel(f"C:\\Users\\kepla\\Downloads\\New_folder3\\complex_pdf_semester3(converted)\\{files2}"))

	#creating an empty dataframe
	excel_merged2 = pd.DataFrame()
	
	for excel_file2 in excel_list2:
		excel_merged2 = excel_merged1.append(excel_file2, ignore_index = True)

	excel_merged1.to_excel("C:\\Users\\kepla\\Downloads\\New_folder3\\complex_pdf_semester3(converted)\\final_converted_complex_result3.xlsx")
	


#function for analysis and report of a docx file

def report_docx():
	#converting docx file to pdf file 
	
	pass

#functions for the events in the file menu

def command_file():
	pass

def command_file1():
	pass

def command_file2():
	pass

def command_file3():
	pass

def command_file4():
	root.quit()

#function for the event in the edit menu

def command_edit():
	pass

def command_edit1():
	pass

def command_edit2():
	pass

def command_edit3():
	pass


def initial_button_function1():
	#prievous file selected
	global root_filename1
	root_filename1 = filedialog.askopenfilename(initialdir="C", title="Select file to Audit", filetypes=(("xlsx file", "*.xlsx"),("all files","*.*")))
	Label22.configure(text="File Selected: "+root_filename1)
	

def initial_button_function():
	#current file selected
	global root_filename
	root_filename = filedialog.askopenfilename(initialdir="C", title="Select file to Audit", filetypes=(("xlsx file", "*.xlsx"),("all files","*.*")))
	Label11.configure(text="File Selected: "+root_filename)

def initial_button_function2():
	#supplementary file selected
	global root_filename2
	root_filename2 = filedialog.askopenfilename(initialdir="C", title="Select file to Audit", filetypes=(("xlsx file", "*.xlsx"),("all files","*.*")))
	Label33.configure(text="File Selected: "+root_filename2)
	
#detection of the extension name of the file

def file_checking():
	#checking whether the file was inputed and whether the file is an xlsx file
	ext = root_filename.split(".")[-1]
	ext1 = root_filename1.split(".")[-1]

	if (options == "YEAR 1" or options == "YEAR 3" or options =="YEAR 2" or options =="YEAR 4") and (options1 == "SECOND SEMESTER" or options1 == "SUPPLEMENTARY SEMESTER"):
		if root_filename == "" or root_filename1 == "":
			messagebox.showwarning("Error","A file(s) was not selected")
		else:
			if ext != "xlsx" or ext1 != "xlsx":
				messagebox.showerror("Error", "A file(s) was not in the excel format")
			else:
				file_checking2_1()
	elif (options == "YEAR 2" or options == "YEAR 3" or options == "YEAR 4") and (options1 == "FIRST SEMESTER"):
		ext2 = root_filename2.split(".")[-1]
		if root_filename == "" or root_filename1 == "" or root_filename2 == "":
			messagebox.showwarning("Error", "A field(s) was not selected")
		else:
			if ext != "xlsx" or ext1 != "xlsx" or ext2 != "xlsx":
				messagebox.showerror("Error", "A file(s) was not is the excel format")
			else:
				file_checking2_2()

def file_checking2_1():
	#checking to see whether the file is comprehensive or summarixed , and checking to see whether file is of the correct year and semester
	#for the first file
		global sentence
		global sentence1
		global shadow_options1
		if options1 == "SUPPLEMENTARY SEMESTER":
			shadow_options1 = "SECOND SEMESTER"
		elif options1 == "SECOND SEMESTER":
			shadow_options1 = "FIRST SEMESTER"


		wb = load_workbook(root_filename)
		wb1 = load_workbook(root_filename1)
		res = len(wb.sheetnames)
		res1 = len(wb1.sheetnames)

		if res != 2 or res1 != 2:
			messagebox.showerror("Error", "The file selected is not comprehensive")

		ws = wb["Sheet1"]
		ws1 = wb1["Sheet1"]
		
		

		row = ws["E8"]
		row1 = ws1["E8"]

		sentence = row.value.split(",")[0]
		sentence1 = row1.value.split(",")[0]
	

		if f"RESULTS OF {options} {options1}" != sentence:
			print(f"RESULTS OF {options} {options1}")
			messagebox.showerror("Error" ,"The file selected is not of the same year or semester")
		elif f"RESULTS OF {options} {shadow_options1}" != sentence1:
			print(f"RESULTS OF {options} {shadow_options1}")
			messagebox.showerror("Error","The file(previous semester) selected is not of the same year or semester")
		else:
			report_xlsx()

		


def file_checking2_2():
		shadow_options1 = "SECOND SEMESTER"
		shadow_options2 = "SUPPLEMENTARY SEMESTER"

		if options == "YEAR 2" and options1 == "FIRST SEMESTER":
	 		shadow_options = "YEAR 1"
		elif options == "YEAR 3" and options1 == "FIRST SEMESTER":
			shadow_options = "YEAR 2"
		elif options == "YEAR 4" and options1 =="FIRST SEMESTER":
			shadow_options = "YEAR 3"

		wb = load_workbook(root_filename)
		wb1 = load_workbook(root_filename1)
		wb2 = load_workbook(root_filename2)
		res = len(wb.sheetnames)
		res1 = len(wb1.sheetnames)
		res2 = len(wb2.sheetnames)

		if res1 != 2 or res != 2 or res2 != 2:
			messagebox.showerror("Error", "File selected is not comprehensive")

		ws = wb["Sheet1"]
		ws1 = wb1["Sheet1"]
		ws2 = wb2["Sheet2"]

		row = ws["E8"]
		row1 = ws1["E8"]
		row2 = ws2["E8"]

		sentence = row.value.split(",")[0]
		sentence1 = row1.value.split(",")[0]
		sentence2 = row2.value.split(",")[0]

		date = row.value.split(",")[1]
		date1 = row.value.split(",")[1]
		date2 = row.value.split(",")[1]


		if f"RESULTS OF {options} {options1}" != sentence:
			messagebox.showerror("Error","File selected is not of the same year or semester")
			print("1")
		elif f"RESULTS OF {shadow_options} {shadow_options1}" != sentence1:
			messagebox.showerror("Error", "File selected is not of the same year or semester")
			print("2")
		elif f"RESULTS OF {shadow_options} {shadow_options2}" != sentence2:
			messagebox.showerror("Error","File selected is not of the same year or semester")
			print("3")
		else:
			report_xlsx1()


#calculating auditing functions for the two file system
def xlsx_audit_function_2FS():
	#checking total number of students tally up
	connection2FS = sqlite3.connect("original_database.db")
	c = connection2FS.cursor()


	#ANALYSING CURRENT FILE 2FS 
	#checking number of students that passed current
	c.execute("SELECT * FROM NUMBER_OF_STUDENTS_PASSED_CURRENT")
	kounter1 = 0
	for i in c.fetchall():
		kounter1 = kounter1 + 1
	final_kounter1 = kounter1 - 2
	print("number of students passed current")
	print(final_kounter1)

	#checking the number of students trailing up to four courses current
	c.execute("SELECT * FROM STUDENTS_TRAILING_UP_TO_FOUR_COURSES_CURRENT")
	kounter2 = 0
	for i in c.fetchall():
		kounter2 = kounter2 + 1
	final_kounter2 = (kounter2 - 3)/2
	print("students trailing up to four courses current")
	print(final_kounter2)

	#checking the number of students trailing more than four courses
	try:
		c.execute("SELECT * FROM STUDENTS_TRAILING_MORE_THAN_FOUR_COURSES_CURRENT")
		kounter3 = 0
		for i in c.fetchall():
			kounter3 = kounter3 + 1
		final_kounter3 = (kounter3 - 3)/2
	except:
		final_kounter3 = 0

	print("students trailing more than four courses current")
	print(final_kounter3)


	#checking for students are have CWA less than 45 current
	try:
		c.execute("SELECT * FROM STUDENTS_WITH_CWA_LESS_THAN_45_CURRENT")
		final_kounter4 = 0
		for i in c.fetchall():
			if type(i[0]) == str:
				final_kounter4 = final_kounter4 + 1
			else:
				continue
		final_kounter4 = final_kounter4 - 1
	except:
		final_kounter4 = 0

	print("students with cwa less than 45 current")
	print(final_kounter4)



	#checking for the students that abandoned the course current

	try:
		c.execute("SELECT * FROM ABANDONED_COURSE_CURRENT")
		kounter5 = 0
		for i in c.fetchall():
			kounter5 = kounter5 + 1
		final_kounter5 = (kounter5 - 3)/2
	except: 
		final_kounter5 = 0 

	print("abandoned course current")
	print(final_kounter5)



	##ANALYSING PREVIOUS FILE 2FS


	#checking the number of students that passed previous

	c.execute("SELECT * FROM NUMBER_OF_STUDENTS_PASSED_PREVIOUS")
	kounter11 = 0
	for i in c.fetchall():
		kounter11 = kounter11 + 1
	final_kounter11 = kounter11 - 2
	print("students that passed all courses previous")
	print(final_kounter11)

	#checking for the number of students that are trailing up to four courses
	c.execute("SELECT * FROM STUDENTS_TRAILING_UP_TO_FOUR_COURSES_PREVIOUS")
	kounter22 = 0
	for i in c.fetchall():
		kounter22 = kounter22 + 1
	final_kounter22 = (kounter22 - 3)/2
	print("Students trailing up to four courses previous")
	print(final_kounter22)

	#checking number of students that are trailing more than four courses previous
	try:
		c.execute("SELECT * FROM STUDENTS_TRAILING_MORE_THAN_FOUR_COURSES_PREVIOUS")
		kounter33 = 0
		for i in c.fetchall():
			kounter33 = kounter33 + 1
		final_kounter33 = (kounter33 - 3)/2
	except:
		final_kounter33 = 0

	print("students trailing more than four courses previous")
	print(final_kounter33)

	#checking for the number of students that have CWA less than 45
	try:
		c.execute("SELECT * FROM STUDENTS_WITH_CWA_LESS_THAN_45_PREVIOUS")
		final_kounter44 = 0
		for i in c.fetchall():
			if type(i[0]) == str:
				final_kounter44 = final_kounter44 + 1
			else:
				continue
		final_kounter44 = final_kounter44 - 1
	except:
		final_kounter44 = 0
	print("Number of students that have CWA less than 45 previous")
	print(final_kounter44)

	#checking for the number of students that have abandoned the courses (if any)
	try:
		c.execute("SELECT * FROM ABANDONED_COURSE_PREVIOUS")
		kounter55 = kounter55 + 1
		for i in c.fetchall():
			kounter55 = kounter55 + 1
		final_kounter55 = (final_kounter55 - 3)/2
	except:
		final_kounter55 = 0

	print("abandoned course previous")
	print(final_kounter55)



	#conditions1(if the current file is not a supplementary file)
	if "SUPPLEMENTARY" not in sentence:
		if final_kounter1 + final_kounter2 + final_kounter3 + final_kounter4 != final_kounter11 + final_kounter22 + final_kounter33 + final_kounter44 + final_kounter55:
			variable = ("REMAKES:The students don't add up")
		else:
			varaible = ("REMAKES:The students add up")
	else:
		if final_kounter2 + final_kounter4 != final_kounter11 + final_kounter22 + final_kounter33 + final_kounter44 + final_kounter55:
			variable = ("REMAKES:The students don't add up")
		else:
			varible = ("REMAKES:The students add up")

	#all variables in this function 



	#creating a docx file to display information on screen
	doc = aw.Document()

	#creating a document builder
	builder = aw.DocumentBuilder(doc)

	#creating a font for the file 
	font = builder.font
	font.size = 15
	font.bold = True
	font.name = "Times New Roman"

	#adding information from analysis to file
	builder.write(f"TOTAL NUMBER OF STUDENT FOR EVERY SECTION(RESULTS OF {options} {options1})")

	#creating a table to insert values
	table = builder.start_table()

	#inserting a cell
	builder.insert_cell()
	table.auto_fit(aw.tables.AutoFitBehavior.AUTO_FIT_TO_CONTENTS)


	#setting format and inserting and adding text
	builder.cell_format.vertical_alignment = aw.tables.CellVerticalAlignment.CENTER
	builder.write("SECTIONS")


	#inserting another cell 

	builder.insert_cell()
	builder.write("TOTAL NUMBER OF STUDENTS FOR EACH SECTION")

	#ending the row 
	builder.end_row()

	#inserting another cell
	builder.insert_cell()

	#formating current row 
	builder.row_format.height = 100
	builder.row_format.height_rule = aw.HeightRule.EXACTLY

	#inserting into current cell

	builder.write("Number of students that passed (current file chosen)")

	#creating cell

	builder.insert_cell()
	builder.write(str(final_kounter1))

	#ending row 
	builder.end_row()

	#inserting another cell
	builder.insert_cell()
	builder.write("Number of students that are trailing up to four courses(current file chosen)")

	builder.insert_cell()
	builder.write(str(final_kounter2))

	builder.end_row()


	builder.insert_cell()
	builder.write("Number of students that are trailing more than four courses(current file chosen)")

	builder.insert_cell()
	builder.write(str(final_kounter3))

	builder.end_row()

	builder.insert_cell()
	builder.write("Number of students that have CWA less than 45(current file selected)")

	builder.insert_cell()
	builder.write(str(final_kounter4))

	builder.end_row()

	builder.insert_cell()
	builder.write("Number of students that abandoned the course (current file selected)")

	builder.insert_cell()
	builder.write(str(final_kounter5))

	builder.end_row()
	builder.end_table()

	
	builder.write(variable)




	



	#creating a different table for previous file selected
	######################################################
	######################################################


	builder1 = aw.DocumentBuilder(doc)

	#creating a font for the file 
	#############################
	font1 = builder1.font
	font1.size = 15
	font1.bold = True
	font1.name = "Times New Roman"

	#adding information from analysis to file
	builder1.write(f"TOTAL NUMBER OF STUDENT FOR EVERY SECTION(RESULTS OF {options} {shadow_options1})")

	#creating a table to insert values
	table1 = builder1.start_table()

	#inserting a cell
	builder1.insert_cell()
	table1.auto_fit(aw.tables.AutoFitBehavior.AUTO_FIT_TO_CONTENTS)


	#setting format and inserting and adding text
	builder1.cell_format.vertical_alignment = aw.tables.CellVerticalAlignment.CENTER
	builder1.write("SECTIONS")


	#inserting another cell 

	builder1.insert_cell()
	builder1.write("TOTAL NUMBER OF STUDENTS FOR EACH SECTION")

	#ending the row 
	builder1.end_row()

	#inserting another cell
	builder1.insert_cell()

	#formating current row 
	builder1.row_format.height = 100
	builder1.row_format.height_rule = aw.HeightRule.EXACTLY

	#inserting into current cell

	builder1.write("Number of students that passed (previous file chosen)")

	#creating cell

	builder1.insert_cell()
	builder1.write(str(final_kounter11))

	#ending row 
	builder1.end_row()

	#inserting another cell
	builder1.insert_cell()
	builder1.write("Number of students that are trailing up to four courses(previous file chosen)")

	builder1.insert_cell()
	builder1.write(str(final_kounter22))

	builder1.end_row()


	builder1.insert_cell()
	builder1.write("Number of students that are trailing more than four courses(previous file chosen)")

	builder1.insert_cell()
	builder1.write(str(final_kounter33))

	builder1.end_row()

	builder1.insert_cell()
	builder1.write("Number of students that have CWA less than 45(previous file selected)")

	builder1.insert_cell()
	builder1.write(str(final_kounter44))

	builder1.end_row()

	builder1.insert_cell()
	builder1.write("Number of students that abandoned the course (previous file selected)")

	builder1.insert_cell()
	builder1.write(str(final_kounter55))


	builder1.end_row()
	builder1.end_table()






	doc.save(f"REPORT FOR {options} {options1} and {shadow_options1}.docx")

	os.startfile(f"REPORT FOR {options} {options1} and {shadow_options1}.docx")

	

























def xlsx_audit_function_3FS():

	connection1 = sqlite3.connect("original_database2.db")
	c1 = connection1.cursor()

	#ANALYSING CURRENT FILE
	#finding the number of students that passed the course current

	c1.execute("SELECT * FROM NUMBER_OF_STUDENTS_PASSED_CURRENT_TFS")
	kounter1TFS = 0
	for i in c1.fetchall():
		kounter1TFS = kounter1TFS + 1
	final_kounter1TFS = kounter1TFS - 2


	print("Number of students that passed the all courses current TFS")
	print(final_kounter1TFS)


	#finding the number of students that trailied up to four courses
	c1.execute("SELECT * FROM STUDENTS_TRAILING_UP_TO_FOUR_COURSES_CURRENT_TFS")
	kounter2TFS = 0
	for i in c1.fetchall():
		kounter2TFS = kounter2TFS + 1
	final_kounter2TFS = (kounter2TFS - 3)/2
	print("students trailing up to four courses current TFS")
	print(final_kounter2TFS)

	#finding the number of students that trailied more than four courses
	try:
		c1.execute("SELECT * FROM STUDENTS_TRAILING_MORE_THAN_FOUR_COURSES_CURRENT_TFS")
		kounter3TFS = 0
		for i in c1.fetchall():
			kounter3TFS = kounter3TFS + 1
		final_kounter3TFS = (kounter3TFS - 3)/2
	except:
		final_kounter3TFS = 0
	print("students trailing more than four courses current tfs")
	print(final_kounter3TFS)



	#finding the number of students that have CWA less than 45 current tfs
	try:
		c1.execute("SELECT * FROM STUDENTS_WITH_CWA_LESS_THAN_45_CURRENT")
		kounter4TFS = 0
		for i in c1.fetchall():
			if type(i[0]) == str:
				kounter4TFS = kounter4TFS + 1
			else:
				continue
		final_kounter4TFS = kounter4TFS - 1
	except:
		final_kounter4TFS = 0
	print("Number of students that have CWA less than 45 previous")
	print(final_kounter4TFS)

	#finding the number of students that have abandoned the course
	try:
		c1.execute("SELECT * FROM ABANDONED_COURSE_CURRENT")
		kounter5TFS = 0
		for i in c1.fetchall():
			kounter5TFS = kounter5TFS + 1
		final_kounter5TFS = (kounter5TFS - 3)/2
	except:
		final_kounter5TFS = 0

	print("abandoned course current")
	print(final_kounter5TFS)




	#ANALYSING PREVIOUS FILE TFS

	#finding the number of students that passed all courses previous TFS
	c1.execute("SELECT * FROM NUMBER_OF_STUDENTS_PASSED_PREVIOUS_TFS")
	kounter11TFS = 0
	for i in c1.fetchall():
		kounter11TFS = kounter11TFS + 1
	final_kounter11TFS = (kounter11TFS - 3)/2
	print("the number of students that passed the previous semester TFS")
	print(final_kounter11TFS)

	#finding the number of students that have up to four trails previous TFS
	c1.execute("SELECT * FROM STUDENTS_TRAILING_UP_TO_FOUR_COURSES_PREVIOUS_TFS")
	kounter22TFS = 0
	for i in c1.fetchall():
		kounter22TFS = kounter22TFS + 1
	final_kounter22TFS = (kounter22TFS - 3)/2
	print("The number of students that are trailing up to four courses previous TFS")
	print(final_kounter22TFS)

	#finding the number of students that are trailing more than four courses
	try:
		c1.execute("SELECT * FROM STUDENTS_TRAILING_MORE_THAN_FOUR_COURSES_PREVIOUS_TFS")
		kounter33TFS = 0
		for i in c1.fetchall():
			kounter33TFS = kounter33TFS + 1
		final_kounter33TFS = (kounter33TFS - 3)/2
	except:
		final_kounter33TFS = 0

	print("students trailing more than four courses previous tfs")
	print(final_kounter33TFS)

	#finding the number of students that have cwa less than 45


	try:
		c1.execute("SELECT * FROM STUDENTS_WITH_CWA_LESS_THAN_45_PREVIOUS")
		kounter44TFS = 0
		for i in c1.fetchall():
			if type(i[0]) == str:
				final_kounter44TFS = final_kounter44TFS + 1
			else:
				continue
		final_kounter44TFS = kounter44TFS - 1
	except:
		final_kounter44TFS = 0
	print("Number of students that have CWA less than 45 previous")
	print(final_kounter44TFS)

	#finding the number of students that have abandoned the courses

	try:
		c1.execute("SELECT * FROM ABANDONED_COURSE_PREVIOUS")
		kounter55TFS = 0
		for i in c1.fetchall():
			kounter55TFS = kounter55TFS + 1
		final_kounter55TFS = (final_kounter55TFS - 3)/2
	except:
		final_kounter55TFS = 0

	print("abandoned course previous")
	print(final_kounter55TFS)



	####
	#ANALYSING SUPPLEMENTARY FILE

	#finding the number of students that passed supplementary

	c1.execute("SELECT * FROM NUMBER_OF_STUDENTS_PASSED_SUPPLEMENTARY_TFS")
	kounter111TFS = 0
	for i in c1.fetchall():
		kounter111TFS = kounter111TFS + 1
	final_kounter111TFS = (kounter111TFS - 3)/2
	print("the number of students that passed the supplementary semester TFS")
	print(final_kounter111TFS)

	#finding the number of students that have up to four trails supplementary TFS
	c1.execute("SELECT * FROM STUDENTS_TRAILING_UP_TO_FOUR_COURSES_SUPPLEMENTARY_TFS")
	kounter222TFS = 0
	for i in c1.fetchall():
		kounter222TFS = kounter222TFS + 1
	final_kounter222TFS = (kounter222TFS - 3)/2
	print("The number of students that are trailing up to four courses supplementary tfs")
	print(final_kounter222TFS)

	#finding the number of students that are trailing more than four courses supplwmentary
	try:
		c1.execute("SELECT * FROM STUDENTS_TRAILING_MORE_THAN_FOUR_COURSES_SUPPLEMENTARY_TFS")
		kounter333TFS = 0
		for i in c1.fetchall():
			kounter333TFS = kounter333TFS + 1
		final_kounter333TFS = (kounter333TFS - 3)/2
		print("students trailing more than four courses supplementary tfs")
		print(final_kounter333TFS)
	except:
		final_kounter333TFS = 0
	print("The number of students that are trailing more than four courses supplementary")
	print(final_kounter333TFS)

	#finding the number of students that have cwa less than 45 supplementary

	try:
		c1.execute("SELECT * FROM STUDENTS_WITH_CWA_LESS_THAN_45_SUPPLEMENTARY")
		kounter444TFS = 0
		for i in c1.fetchall():
			if type(i[0]) == str:
				kounter444TFS = kounter444TFS + 1
			else:
				continue
		final_kounter444TFS = kounter444TFS - 1
	except:
		final_kounter444TFS = 0
	print("Number of students that have CWA less than 45 previous")
	print(final_kounter444TFS)

	#finding the number of students that have abandoned the courses supplementary

	try:
		c1.execute("SELECT * FROM ABANDONED_COURSE_SUPPLEMENTARY")
		kounter555TFS = 0
		for i in c1.fetchall():
			kounter555TFS = kounter555TFS + 1
		final_kounter555TFS = (kounter555TFS - 3)/2
	except:
		final_kounter555TFS = 0

	print("abandoned course previous")
	print(final_kounter555TFS)


	#CONDITIONS NEEDED
	if final_kounter1TFS + final_kounter2TFS + final_kounter3TFS + final_kounter4TFS + final_kounter5TF == (final_kounter11TFS + final_kounter22TFS + final_kounter33TFS + final_kounter44TFS) + (final_kounter111TFS + final_kounter222TFS + final_kounter333TFS + final_kounter444TFS):
		print("The students add up")
	else:
		print("The students don't add up")

	#final_kounter1TFS (Number of students that passed current)
	#final_kounter2TFS(Number of students that are trailing up to four courses current)
	#final_kounter3TFS(Number of students that are trailing more than four courses current)
	#fianl_kounter4TFS(Number of students that have CWA less than 45 current)
	#fianl_kounter5TFS(Number of students that abandoned the course current)
	#fianl_kounter11TFS(Number of students that passed previous)
	#fianl_kounter22TFS(Number of students that are trailing up to four courses previous)
	#fianl_kounter33TFS(Number of students that are trailing more than four courses previous)
	#fianl_kounter44TFS(Number of students that have CWA less than 45 previou)
	#fianl_kounter55TFS(Number of students that hava abandoned the course previous)
	#fianl_kounter111TFS(Number of students that have passed supplementry)
	#fianl_kounter222TFS(Number of students that are trailing up to four courses supplementry)
	#fianl_kounter333TFS(Number of students that are trailing more than four courses supplementry)
	#fianl_kounter444TFS(Number of students that have CWA less than 45)
	#fianl_kounter555TFS(Number of students that have abandoned the course)
	


def paranthesis_remover(value):
	convert = list(value)
	convert.remove("(")
	convert.remove(")")
	interger_convert = int(convert[0])
	return interger_convert



def auditing_function2_checking_credit_registered2FS():
	#CHECKING THE TOTAL NUMBER OF CREDIT HOURS FOR SEMESTER
	connection2 = sqlite3.connect("original_database.db")
	c3 = connection2.cursor()
	c3.execute("SELECT * FROM NUMBER_OF_STUDENTS_PASSED_CURRENT")
	default_list = []
	functional_list = [] #list i will be working with storing all the credit hours for the current courses that are been taken
	if "SUPPLEMENTARY" in sentence:
		#FINDING THE APPROPRIATE NUMBER OF REGISTERED CREDIT HOURS (IN CASE CURRENT IS SUPPLEMENTARY)
		total_credit_hours_supplementary = 0
		for o in c3.fetchmany(3):
			for i in o[3]:
				for x in range(15, 15 + len(o)):
					if i[x] != None:
						total_credit_hours_supplementary = total_credit_hours_supplementary + i[x]
					else:
						continue
		print(total_credit_hours_supplementary)

				


		
	else:
		for i in c3.fetchmany(2):
			default_list.append(i)
		total_credit_hours = 0
		functional_list = list(default_list[1])
		for x in range(14, len(functional_list)):
			if functional_list[x] != None:
				number = paranthesis_remover(functional_list[x])
				total_credit_hours = total_credit_hours + number
			else:
		
				continue
	#finding whether all students have the right number of credit hours registerd and calculated
	#current 2FS
	for a in c3.fetchall():
		if int(a[4]) == total_credit_hours:
			print(a[3] + " had accurate credit registered")
		else:
			print(a[3] + " had inaccurate credit registered")












			



	
		
		




































def next():
	global options,options1
	
	#checking for the options chosen (the year and the semester)
	options = clicked.get()
	options1 = clicked1.get()
	global Label11, Label22, Label33, button_on_initial1, button_on_audit, button_on_initial, button_on_audit, Beginning, Label33, Beginning2, button_back
	button_back = Button(frame, text="back" , command=destroyer)

	if (options == "YEAR 1") and (options1 == "FIRST SEMESTER"):
		messagebox.showerror("Error", "Year1 semster1 cannot be compared with any other file")
	elif (options == "YEAR 1" or options == "YEAR 3" or options =="YEAR 2" or options =="YEAR 4") and (options1 == "SECOND SEMESTER" or options1 == "SUPPLEMENTARY SEMESTER"):
		'''top = Toplevel()
		top.geometry("800x300")
		top.title("Audit")'''
		dropbox.destroy()
		dropbox1.destroy()
		button_next.destroy()
		button_back.grid(row=14, column=5, padx=15, pady=15)

	
		Beginning = Label(frame, text="Select current semester file")
		Beginning.grid(row=1, column=3)

		Label11 = Label(frame, text="No file selected")
		Label11.grid(row=2, column=4)

		button_on_initial = Button(frame, text="Browse" , command=initial_button_function)
		button_on_initial.grid(row=2 , column=1, padx=15)


		Beginning2 = Label(frame, text="Select previous semester file")
		Beginning2.grid(row=3, column=3)

		Label22 = Label(frame, text="No file selected")
		Label22.grid(row=4, column=4)

		button_on_initial1 = Button(frame, text="Browse", command=initial_button_function1)
		button_on_initial1.grid(row=4, column=1)


		button_on_audit = Button(frame, text="Audit", command=file_checking)
		button_on_audit.grid(row=8, column=2,padx =15)
	
	elif (options == "YEAR 2" or options == "YEAR 3" or options == "YEAR 4") and (options1 == "FIRST SEMESTER"):
		'''top1 = Toplevel()
		top1.geometry("800x300")
		top1.title("Audit")'''
		dropbox.destroy()
		dropbox1.destroy()
		button_next.destroy()
		button_back.grid(row=14, column=5, padx=15, pady=15)

		Beginning = Label(frame, text="Select current semester file")
		Beginning.grid(row=1, column=3)

		Label11 = Label(frame, text="No file selected")
		Label11.grid(row=2, column=4)

		button_on_initial = Button(frame, text="Browse", command=initial_button_function)
		button_on_initial.grid(row=2, column=1, padx=15)


		Beginning2 = Label(frame, text="Select previous semester file")
		Beginning2.grid(row=3, column=3)

		Label22 = Label(frame, text="No file selected")
		Label22.grid(row=4, column=4)

		button_on_initial1 = Button(frame, text="Browse", command=initial_button_function1)
		button_on_initial1.grid(row=4, column=1)


		Beginning2 = Label(frame, text="Select previous semester3 file")
		Beginning2.grid(row=5, column=3)

		Label33 = Label(frame, text="No file selected")
		Label33.grid(row=6, column=4)

		button_on_initial2 = Button(frame, text="Browse", command=initial_button_function2)
		button_on_initial2.grid(row=6, column=1)


		button_on_audit = Button(frame, text="Audit", command=file_checking)
		button_on_audit.grid(row=8, column=2,padx =15)

		li = [Label11, Label22, Label33, button_on_initial1, button_on_audit, button_on_initial, button_on_audit, Beginning, Label33, Beginning2,button_on_initial2]

def destroyer():
	frame.destroy()
	back()

#where a is the header number	
def numberforowscalculated(a):
	numberofrows1 = 3
	for i in range(a,3000):
		if ws.cell(row=i , column=4).value != None:
			numberofrows1 = numberofrows1 + 1
		else:
			break

	return numberofrows1;
#current
def students_trailing_up_to_four_courses():
	rownumberTU4 = 0
	for i in range(header1_1 + 3, len(df2)):
		if pd.notnull(df2.iat[i, 1]):
			rownumberTU4 = rownumberTU4 + 1
		else:
			break
	new_df_2 = df2.iloc[range(header1_1 + 1, header1_1 + 2 + rownumberTU4 + 1),:]
	del new_df_2["Unnamed: 4"]
	del new_df_2["Unnamed: 18"]
	del new_df_2["Unnamed: 12"]
	new_df_2.to_sql("STUDENTS_TRAILING_UP_TO_FOUR_COURSES_CURRENT", con = engine, if_exists = "replace", index = False)
#current
def students_trailing_more_than_four_courses():
	rownumberMT4 = 0
	for i in range(header2_2 + 3, len(df2)):
		if pd.notnull(df2.iat[i, 1]):
			rownumberMT4 = rownumberMT4 + 1
		else:
			break
	new_df_2 = df2.iloc[range(header2_2 + 1, header2_2 + 2 + rownumberMT4 + 1),:]
	del new_df_2["Unnamed: 4"]
	del new_df_2["Unnamed: 18"]
	del new_df_2["Unnamed: 12"]
	new_df_2.to_sql("STUDENTS_TRAILING_MORE_THAN_FOUR_COURSES_CURRENT", con = engine, if_exists = "replace", index = False)
#current
def students_with_CWA_less_than_45():
	rownumberCWA45 = 0
	for i in range(header3_3 + 3, len(df2)):
		if pd.notnull(df2.iat[i, 1]):
			rownumberCWA45 = rownumberCWA45 + 1
		else:
			break
	new_df_2 = df2.iloc[range(header3_3 + 1, header3_3 + 2 + rownumberCWA45 + 1),:]
	del new_df_2["Unnamed: 4"]
	del new_df_2["Unnamed: 18"]
	del new_df_2["Unnamed: 12"]
	new_df_2.to_sql("STUDENTS_WITH_CWA_LESS_THAN_45_CURRENT", con = engine, if_exists = "replace", index = False)
#current
def Abandoned():
	rownumberAboned = 0
	for i in range(header4_4 + 3, len(df2)):
		if pd.notnull(df2.iat[i, 1]):
			rownumberAboned = rownumberAboned + 1
		else:
			break
	new_df_2 = df2.iloc[range(header4_4 + 1, header4_4 + 2 + rownumberAboned + 1),:]
	del new_df_2["Unnamed: 4"]
	del new_df_2["Unnamed: 18"]
	del new_df_2["Unnamed: 12"]
	new_df_2.to_sql("ABANDONED_COURSE_CURRENT", con = engine, if_exists = "replace", index = False)








#previous




def students_trailing_up_to_four_coursesp():
	rownumberTU4p = 0
	for i in range(header1_11 + 3, len(df3)):
		if pd.notnull(df3.iat[i, 1]):
			rownumberTU4p = rownumberTU4p + 1
		else:
			break
	new_df_2p = df3.iloc[range(header1_11 + 1, header1_11 + 2 + rownumberTU4p + 1),:]
	del new_df_2p["Unnamed: 4"]
	del new_df_2p["Unnamed: 18"]
	del new_df_2p["Unnamed: 12"]
	new_df_2p.to_sql("STUDENTS_TRAILING_UP_TO_FOUR_COURSES_PREVIOUS", con = engine, if_exists = "replace", index = False)
#current
def students_trailing_more_than_four_coursesp():
	rownumberMT4p = 0
	for i in range(header2_22 + 3, len(df3)):
		if pd.notnull(df3.iat[i, 1]):
			rownumberMT4p = rownumberMT4p + 1
		else:
			break
	new_df_2p = df3.iloc[range(header2_22 + 1, header2_22 + 2 + rownumberMT4p + 1),:]
	del new_df_2p["Unnamed: 4"]
	del new_df_2p["Unnamed: 18"]
	del new_df_2p["Unnamed: 12"]
	new_df_2p.to_sql("STUDENTS_TRAILING_MORE_THAN_FOUR_COURSES_PREVIOUS", con = engine, if_exists = "replace", index = False)
#current
def students_with_CWA_less_than_45p():
	print("hello")
	rownumberCWA45p = 0
	for i in range(header3_33 + 3, len(df3)):
		if pd.notnull(df3.iat[i, 1]):
			rownumberCWA45p = rownumberCWA45p + 1
		else:
			break
	new_df_2p = df3.iloc[range(header3_33 + 1, header3_33 + 2 + rownumberCWA45p + 1),:]
	del new_df_2p["Unnamed: 4"]
	del new_df_2p["Unnamed: 18"]
	del new_df_2p["Unnamed: 12"]
	new_df_2p.to_sql("STUDENTS_WITH_CWA_LESS_THAN_45_PREVIOUS", con = engine, if_exists = "replace", index = False)
#current
def Abandonedp():
	rownumberAbonedp = 0
	for i in range(header4_44 + 3, len(df3)):
		if pd.notnull(df3.iat[i, 1]):
			rownumberAbonedp = rownumberAbonedp + 1
		else:
			break
	new_df_2p = df3.iloc[range(header4_44 + 1, header4_44 + 2 + rownumberAbonedp + 1),:]
	del new_df_2p["Unnamed: 4"]
	del new_df_2p["Unnamed: 18"]
	del new_df_2p["Unnamed: 12"]
	new_df_2p.to_sql("ABANDONED_COURSE_PREVIOUS", con = engine, if_exists = "replace", index = False)



def back(): 


	global dropbox, dropbox1, clicked1, clicked, frame, button_next
	frame = LabelFrame(root, text="Please select the required feild" , padx=5, pady=5)
	frame.pack(padx=20, pady=20)

	clicked = StringVar()
	clicked.set("YEAR 1")

	clicked1 = StringVar()
	clicked1.set("FIRST SEMESTER")

	dropbox = OptionMenu(frame,clicked,"YEAR 1", "YEAR 2", "YEAR 3", "YEAR 4",)
	dropbox.grid(row=4, column=5, padx=155, pady=15) 

	dropbox1 = OptionMenu(frame, clicked1, "FIRST SEMESTER","SECOND SEMESTER","SUPPLEMENTARY SEMESTER")
	dropbox1.grid(row=10, column=5, padx=15, pady=15)

	button_next = Button(frame, text="Next" , command=next)
	button_next.grid(row=14, column=5, padx=15, pady=15)


root_filename = ""
root_filename1 = ""
root_filename2 = ""
back()




















"""Beginning = Label(root, text="Select semester1 file")
Beginning.grid(row=1, column=3)

#creating a button for browse

button_on_initial = Button(root, text="Browse" , command=initial_button_function)
button_on_initial.grid(row=2 , column=2, padx=15)



Label11 = Label(root, text="No file selected")
Label11.grid(row=2, column=4)

Beginning = Label(root, text="Select semester2 file")
Beginning.grid(row=4, column=3)

button_on_initial1 = Button(root, text="Browse", command=initial_button_function1)
button_on_initial1.grid(row=6, column=2, padx=15)




Label22 = Label(root, text="No file selected")
Label22.grid(row=6 , column=4)


button_on_audit = Button(root, text="Audit", command=extention_detection)
button_on_audit.grid(row=9, column=3,padx =15)

button_on_supplementary = Button(root, text="Browse", command=initial_button_function2)
button_on_supplementary.grid(row=8, column=2, padx=15)



Label33 = Label(root, text="No file selected")
Label33.grid(row=8, column=4)

Label44 = Label(root, text = "Select supplementary file")
Label44.grid(row=7, column=3)


#creating a menu

my_menu = Menu(root)
root.config(menu=my_menu)

#creating the cascading items for the file menu

file_menu = Menu(my_menu)
my_menu.add_cascade(label="File",menu=file_menu)
file_menu.add_command(label="New", command=command_file)
file_menu.add_command(label="def1", command=command_file1)
file_menu.add_command(label="def2", command=command_file2)
file_menu.add_command(label="def3", command=command_file3)
file_menu.add_separator()
file_menu.add_command(label="Exit", command=command_file4)

#creating the cascading items for the edit menu 

edit_menu = Menu(my_menu)
my_menu.add_cascade(label="Edit", menu=edit_menu)
edit_menu.add_command(labe="def11", command=command_edit)
edit_menu.add_command(labe="def12", command=command_edit1)
edit_menu.add_command(labe="def13", command=command_edit2)
edit_menu.add_command(labe="def14", command=command_edit3)"""

root.mainloop()





